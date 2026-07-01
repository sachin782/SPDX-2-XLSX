"""Excel workbook writer for SPDX data."""

import pathlib
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils import logger

# Styling constants
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="2E4057", end_color="2E4057")
CELL_FONT = Font(name="Arial", size=10)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
MAX_COL_WIDTH = 60
MIN_COL_WIDTH = 12


def _write_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    rows: list[dict],
) -> None:
    """Write rows to a worksheet with formatting.

    Args:
        ws: Target worksheet.
        rows: List of dicts; keys become column headers.
    """
    if not rows:
        return

    headers = list(rows[0].keys())

    # Write header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-adjust column widths
    col_widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                # Account for multiline text
                lines = str(cell.value).split("\n")
                longest = max(len(line) for line in lines)
                col_widths[cell.column] = max(col_widths.get(cell.column, 0), longest)

    for col_idx, width in col_widths.items():
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(
            MIN_COL_WIDTH, min(width + 2, MAX_COL_WIDTH)
        )

    # Set row height for data rows to accommodate wrapped text
    ws.row_dimensions[1].height = 20
    logger.debug("Sheet '%s' written with %d data rows.", ws.title, len(rows))


def build_workbook(sections: dict[str, list[dict]]) -> openpyxl.Workbook:
    """Create an openpyxl workbook from parsed SPDX sections.

    Args:
        sections: Mapping of sheet name → list of row dicts.

    Returns:
        Populated Workbook object (not yet saved).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    for sheet_name, rows in sections.items():
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, rows)
        logger.info("Created sheet '%s' with %d rows.", sheet_name, len(rows))

    return wb


def save_workbook(wb: openpyxl.Workbook, output_path: pathlib.Path) -> None:
    """Save a workbook to disk.

    Args:
        wb: The workbook to save.
        output_path: Full destination path including filename.

    Raises:
        OSError: If the file cannot be written.
    """
    wb.save(output_path)
    logger.info("Workbook saved to %s", output_path)
